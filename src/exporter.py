# src/exporter.py
import csv
import openpyxl
from fpdf import FPDF
import io

def generate_csv(kontakte, eigenschaften):
    """Erstellt eine CSV-Datei im Speicher."""
    output = io.StringIO()
    writer = csv.writer(output)
    
    headers = [e['name'] for e in eigenschaften]
    writer.writerow(headers)
    
    for kontakt in kontakte:
        row = [str(kontakt['daten'].get(h, '')) for h in headers]
        writer.writerow(row)
        
    output.seek(0)
    return output.getvalue()

def generate_xlsx(kontakte, eigenschaften):
    """Erstellt eine Excel-Datei (.xlsx) im Speicher."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    
    headers = [e['name'] for e in eigenschaften]
    sheet.append(headers)
    
    for kontakt in kontakte:
        row = [kontakt['daten'].get(h, '') for h in headers]
        sheet.append(row)
        
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()

def generate_pdf(kontakte, eigenschaften, vorlage_name):
    """Erstellt eine PDF-Datei im Speicher für den Druck."""
    pdf = FPDF(orientation='L') # Querformat für mehr Spalten
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    
    pdf.cell(0, 10, txt=f"Kontakt-Export: {vorlage_name}", ln=True, align='C')
    
    headers = [e['name'] for e in eigenschaften]
    
    # Automatische Spaltenbreite (vereinfacht)
    num_columns = len(headers)
    col_width = pdf.w / (num_columns + 0.1) if num_columns > 0 else 10

    # Header
    pdf.set_font("Arial", 'B', 8)
    for header in headers:
        pdf.cell(col_width, 10, header, border=1, align='C')
    pdf.ln()
    
    # Datenzeilen
    pdf.set_font("Arial", '', 8)
    for kontakt in kontakte:
        for header in headers:
            value = str(kontakt['daten'].get(header, ''))
            pdf.cell(col_width, 10, value, border=1)
        pdf.ln()

    return pdf.output(dest='S').encode('latin-1')